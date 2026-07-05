"""
Анализ рисков ОСФ: Рейтинг РУСАДА × Прогнозная модель нарушений.
Объект анализа — виды спорта (ОСФ), без регионов.

Логика агрегации рисков — точно по ipynb модели:
  sort_by_zone (🔴→🟠→🟢) → внутри по proba убыв. → groupby("Вид спорта").first()

Ключевые принципы этой версии (исправления прошлых ошибок):
  1. Рейтинг читается НАПРЯМУЮ из PDF (pdfplumber), без промежуточного JSON.
  2. Итоговая таблица содержит ВСЕ федерации из рейтинга (по умолчанию 142),
     а не только те, что нашлись в прогнозной модели.
  3. Критерии рейтинга НЕ бинарны: встречается значение 2. «Выполнен» = значение >= 1.
  4. Матчинг вид спорта ↔ ОСФ идёт по нормализованным ТОКЕНАМ + словарю алиасов,
     а НЕ по наивному вхождению подстроки (иначе "бокс" ошибочно ловит "кикбоксинг").
  5. ОСФ без матча в модели попадают в таблицу со статусом «нет данных модели»
     и НЕ считаются рисковыми по умолчанию.
  6. Ведётся аудит матчинга (исходное имя из модели, тип матчинга, уверенность).

ИСПОЛЬЗОВАНИЕ:
  1) Обновите PDF_PATH, IPYNB_PATH, OUT_DIR ниже.
  2) sudo pip3 install pdfplumber   (если ещё не установлен)
  3) python build_report.py
"""

import os
import re
import json

# ── ПУТИ К ФАЙЛАМ (обновить перед запуском) ───────────────────────────────
PDF_PATH   = "/tmp/Reiting-OSF.pdf"                 # PDF «Рейтинг ОСФ 2025» с rusada.ru
IPYNB_PATH = "/home/ubuntu/upload/2_predict.ipynb"  # ipynb прогнозной модели рисков
OUT_DIR    = "/home/ubuntu/"
# ─────────────────────────────────────────────────────────────────────────

import pandas as pd
import numpy as np
import pdfplumber
import plotly.graph_objects as go
import plotly.express as px
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── СТИЛЬ ──────────────────────────────────────────────────────────────────
INK   = "#0F2D52"
SUB   = "#7C8DA6"
GRID  = "#EEF2F7"
FONT  = "Inter, Segoe UI, Roboto, Arial, sans-serif"

# Текстовые коды зон (для логики, сортировки, фильтрации)
ZONE_CODE_RED    = "RED"
ZONE_CODE_ORANGE = "ORANGE"
ZONE_CODE_GREEN  = "GREEN"
ZONE_CODE_NODATA = "NO_DATA"

# Эмодзи — только декор в Excel и дашборде
ZONE_EMOJI = {"RED": "🔴", "ORANGE": "🟠", "GREEN": "🟢", "NO_DATA": "⚪"}

# Для бар-графика и Excel (3 зоны):
ZONE_COLORS = {"RED": "#DC2626", "ORANGE": "#F59E0B", "GREEN": "#10B981", "NO_DATA": "#94A3B8"}

# Для матрицы квадрантов scatter (4 приоритета):
QUADRANT_COLORS = {
    "Приоритет 1 (низкий рейтинг + систематичность)": "#DC2626",
    "Приоритет 2 (высокий рейтинг + систематичность)": "#F59E0B",
    "Приоритет 3 (низкий рейтинг + нет систематичности)": "#3B82F6",
    "Приоритет 4 (высокий рейтинг + нет систематичности)": "#10B981",
}
# Правило: scatter красится по QUADRANT_COLORS, бар и Excel — по ZONE_COLORS. Не смешивать.

# Обратная совместимость для старых частей кода
NO_DATA_ZONE = "⚪ НЕТ ДАННЫХ"
ZONE_COLORS_FULL = {"🔴 КРАСНАЯ": "#DC2626", "🟠 ОРАНЖЕВАЯ": "#F59E0B", "🟢 ЗЕЛЁНАЯ": "#10B981", NO_DATA_ZONE: "#94A3B8"}
QUAD_COLORS = {
    k: v for k, v in QUADRANT_COLORS.items()
}

CRITERIA = ["Стратегия", "План-график", "Регионы", "Сайт",
            "Семинар", "Соглашение", "Допуск", "Мониторинг", "Инфо"]

RATING_THRESHOLD = 80  # порог «высокий рейтинг» (баллы РУСАДА)


def _theme(fig, title, subtitle=""):
    sub_html = f"<br><span style='font-size:12px;color:{SUB}'>{subtitle}</span>" if subtitle else ""
    fig.update_layout(
        template="plotly_white",
        font=dict(family=FONT, color=INK, size=13),
        title=dict(text=f"<b>{title}</b>{sub_html}", x=0.01, xanchor="left", y=0.97,
                   font=dict(size=18)),
        margin=dict(l=10, r=24, t=84, b=24),
        paper_bgcolor="white", plot_bgcolor="white", legend_title_text="",
    )
    fig.update_xaxes(showgrid=False, zeroline=False)
    fig.update_yaxes(showgrid=False, zeroline=False)
    return fig


# ══════════════════════════════════════════════════════════════════
# ШАГ 1. ЗАГРУЗКА РЕЙТИНГА ОСФ НАПРЯМУЮ ИЗ PDF (все федерации)
# ══════════════════════════════════════════════════════════════════

def load_osf_rating(pdf_path):
    """Извлекает все строки рейтинга из PDF.

    Структура строки таблицы:
      [Федерация, Стратегия, План-график, Регионы, Сайт, Семинар,
       Соглашение, Допуск, Мониторинг, Инфо, СУММА, МЕСТО, (иногда пустая ячейка)]

    Некоторые строки приходят с 13 колонками (лишняя пустая ячейка в конце) —
    это не влияет на разбор, т.к. критерии всегда в позициях 1..9,
    СУММА в позиции 10, МЕСТО в позиции 11.
    """
    raw = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                raw.extend(table)

    rows = []
    for r in raw:
        if not r or r[0] is None:
            continue
        name = str(r[0]).replace("\n", " ").strip()
        name = re.sub(r"\s+", " ", name)
        # Пропускаем строку-заголовок (повторяется на каждой странице)
        if name in ("", "Федерация") or name.lower().startswith("федерация\n"):
            continue

        # Критерии: позиции 1..9. «Выполнен» = значение >= 1 (встречается и 2).
        vals = []
        for v in r[1:10]:
            try:
                vals.append(int(str(v).strip()))
            except Exception:
                vals.append(0)

        try:
            score = int(str(r[10]).strip())
        except Exception:
            score = None

        try:
            place = int(str(r[11]).strip())
        except Exception:
            place = None

        # Валидная строка данных = есть имя и распарсенный балл
        if name and score is not None:
            rows.append([name] + vals + [score, place])

    df = pd.DataFrame(rows, columns=["ОСФ"] + CRITERIA + ["Баллы", "Место"])
    df = df.drop_duplicates(subset=["ОСФ"]).reset_index(drop=True)

    def done(row):
        return "; ".join(c for c in CRITERIA if row[c] >= 1)

    def not_done(row):
        return "; ".join(c for c in CRITERIA if row[c] < 1)

    df["Выполненные критерии"]   = df.apply(done, axis=1)
    df["Невыполненные критерии"] = df.apply(not_done, axis=1)
    df["_key"] = df["ОСФ"].apply(_norm)
    df["_tokens"] = df["ОСФ"].apply(_sport_tokens)
    return df


# ══════════════════════════════════════════════════════════════════
# ШАГ 2. ИЗВЛЕЧЕНИЕ РИСКОВ ИЗ IPYNB
# ══════════════════════════════════════════════════════════════════

ZONE_ORDER = {"🔴 КРАСНАЯ": 0, "🟠 ОРАНЖЕВАЯ": 1, "🟢 ЗЕЛЁНАЯ": 2,
              "RED": 0, "ORANGE": 1, "GREEN": 2, "NO_DATA": 3}


def load_risks_from_ipynb(ipynb_path):
    with open(ipynb_path, "r", encoding="utf-8") as f:
        nb = json.load(f)

    records = []
    for cell in nb.get("cells", []):
        for output in cell.get("outputs", []):
            pdata = output.get("data", {}).get("application/vnd.plotly.v1+json", {})
            if not pdata:
                continue
            for trace in pdata.get("data", []):
                y_vals = trace.get("y", [])
                cdata  = trace.get("customdata", [])
                txt    = trace.get("text", [])
                if not y_vals or not cdata:
                    continue
                for i, label in enumerate(y_vals):
                    try:
                        sport = str(label).split(" — ")[0].strip()
                        cd    = cdata[i] if i < len(cdata) else []
                        reason = cd[0] if len(cd) > 0 else ""
                        zone   = cd[1] if len(cd) > 1 else "🟢 ЗЕЛЁНАЯ"
                        proba  = float(txt[i]) if i < len(txt) else 0.0
                        records.append({
                            "Вид спорта": sport,
                            "зона_риска": zone,
                            "proba": proba,
                            "причина": reason,
                        })
                    except Exception:
                        continue

    df = pd.DataFrame(records)
    if df.empty:
        return df

    df["_z"] = df["зона_риска"].map(ZONE_ORDER).fillna(3)
    df = df.sort_values(["_z", "proba"], ascending=[True, False]).drop(columns="_z")
    agg = df.groupby("Вид спорта", as_index=False).first()
    agg["_key"] = agg["Вид спорта"].apply(_norm)
    agg["_tokens"] = agg["Вид спорта"].apply(_sport_tokens)
    return agg


# ══════════════════════════════════════════════════════════════════
# ШАГ 3. МАТЧИНГ ОСФ ↔ ВИД СПОРТА (по токенам + алиасы)
# ══════════════════════════════════════════════════════════════════

# Слова, не несущие информацию о виде спорта.
STOPWORDS = {
    "федерация", "всероссийская", "всероссийское", "российская", "российский",
    "союз", "ассоциация", "объединение", "объединенная", "национальная",
    "национальный", "общероссийская", "спортивная", "спортивно", "спорта",
    "спорт", "россии", "россия", "лиц", "с", "и", "на", "в", "по", "совет",
    "офсоо", "федерации", "видов", "года", "деятельности", "антидопинговой",
    "кинологической", "системе", "заболеванием",
}

# Алиасы: ключ = нормализованное имя вида спорта из МОДЕЛИ,
# значение = фрагмент имени ОСФ, который должен входить в нормализованное
# имя ОСФ. Значения прогоняются через _norm() при сравнении, поэтому можно писать естественно (й/ё).
# Заполняется по мере обнаружения расхождений формулировок «модель ↔ рейтинг».
ALIASES = {
    "плавание": "водных видов спорта",
    "прыжки в воду": "водных видов спорта",
    "водное поло": "водных видов спорта",
    "синхронное плавание": "водных видов спорта",
    "легкая атлетика": "легкой атлетики",
    "велоспорт": "велосипедного спорта",
    "велоспорт шоссе": "велосипедного спорта",
    "велоспорт трек": "велосипедного спорта",
    "маунтинбайк": "велосипедного спорта",
    "бмх": "велосипедного спорта",
    "гребля": "гребного спорта",
    "гребной спорт": "гребного спорта",
    "академическая гребля": "гребного спорта",
    "гребля на байдарках и каноэ": "каноэ",
    "гребной слалом": "каноэ",
    "тяжелая атлетика": "тяжелой атлетики",
    "конькобежный спорт": "конькобежцев",
    "коньки": "конькобежцев",
    "шорт-трек": "конькобежцев",
    "фигурное катание": "фигурного катания",
    "лыжные гонки": "лыжных гонок",
    "лыжное двоеборье": "прыжков на лыжах",
    "прыжки на лыжах с трамплина": "прыжков на лыжах",
    "биатлон": "биатлонистов",
    "футбол": "футбольный союз",
    "мини-футбол": "футбольный союз",
    "баскетбол": "баскетбола",
    "гандбол": "гандбола",
    "волейбол": "волейбола",
    "пляжный волейбол": "волейбола",
    "борьба": "спортивной борьбы",
    "вольная борьба": "спортивной борьбы",
    "греко-римская борьба": "спортивной борьбы",
    "самбо": "самбо",
    "дзюдо": "дзюдо",
    "бокс": "бокса",
    "кикбоксинг": "кикбоксинга",
    "тхэквондо": "тхэквондо",
    "каратэ": "каратэ",
    "муайтай": "муайтай",
    "тайский бокс": "муайтай",
    "пауэрлифтинг": "пауэрлифтинга",
    "бодибилдинг": "бодибилдинга",
    "художественная гимнастика": "гимнастики",
    "спортивная гимнастика": "гимнастики",
    "прыжки на батуте": "гимнастики",
    "теннис": "тенниса",
    "настольный теннис": "настольного тенниса",
    "гольф": "гольфа",
    "регби": "регби",
    "хоккей": "хоккея россии",
    "хоккей на траве": "хоккея на траве",
    "хоккей с мячом": "хоккея с мячом",
    "триатлон": "триатлона",
    "современное пятиборье": "современного пятиборья",
    "стрельба из лука": "стрельбы из лука",
    "пулевая стрельба": "пулевой стрельбы",
    "стендовая стрельба": "пулевой стрельбы",
    "фехтование": "фехтования",
    "конный спорт": "конного спорта",
    "бадминтон": "бадминтона",
    "скалолазание": "скалолазания",
    "сноуборд": "сноуборда",
    "фристайл": "фристайла",
    "прыжки в воду вышка": "водных видов спорта",
}


def _norm(s):
    """Нормализация: нижний регистр, ё→е, схлопывание пробелов и дефисов, чистка кавычек."""
    s = str(s).lower().replace("ё", "е")
    s = s.replace("«", " ").replace("»", " ").replace('"', " ").replace("'", " ")
    s = re.sub(r"[\-–—]", " ", s)
    s = re.sub(r"\([^)]*\)", " ", s)   # убираем скобочные пояснения
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _sport_tokens(name):
    """Значимые токены вида спорта (без стоп-слов, длина > 2)."""
    return {w for w in _norm(name).split() if w not in STOPWORDS and len(w) > 2}


def match_sport_to_osf(sport_name, osf_df):
    """Возвращает (osf_row, match_type, confidence) или (None, 'нет матча', 0)."""
    sn = _norm(sport_name)
    s_tokens = _sport_tokens(sport_name)

    # 1) Явный алиас модель→ОСФ (target прогоняется через _norm — снимает расхождения й/ё)
    if sn in ALIASES:
        target = _norm(ALIASES[sn])
        for _, row in osf_df.iterrows():
            if target in row["_key"]:
                return row, "алиас", 1.0

    # 2) Точное совпадение значимых токенов (пересечение непустое и полное с одной из сторон)
    best, best_score = None, 0.0
    for _, row in osf_df.iterrows():
        o_tokens = row["_tokens"]
        if not s_tokens or not o_tokens:
            continue
        inter = s_tokens & o_tokens
        if not inter:
            continue
        # доля общих токенов относительно вида спорта из модели
        score = len(inter) / len(s_tokens)
        # требуем полного покрытия токенов модели (чтобы "бокс" != "кикбоксинг")
        if s_tokens.issubset(o_tokens) and score > best_score:
            best, best_score = row, score
    if best is not None:
        return best, "токены", round(best_score, 2)

    # 3) Ослабленное правило: хотя бы один значимый токен модели совпал ЦЕЛИКОМ
    for _, row in osf_df.iterrows():
        if s_tokens & row["_tokens"]:
            return row, "частичный токен", 0.5

    return None, "нет матча", 0.0


# ══════════════════════════════════════════════════════════════════
# ШАГ 4. СБОРКА ИТОГОВОЙ ТАБЛИЦЫ ПО ВСЕМ ОСФ
# ══════════════════════════════════════════════════════════════════

def build_final_table(osf_df, risks_df):
    # Индекс рисков по ОСФ: для каждой ОСФ ищем сматченный вид спорта
    osf_to_risk = {}
    audit = []

    if not risks_df.empty:
        for _, risk in risks_df.iterrows():
            osf_row, mtype, conf = match_sport_to_osf(risk["Вид спорта"], osf_df)
            audit.append({
                "Вид спорта (модель)": risk["Вид спорта"],
                "ОСФ (рейтинг)": osf_row["ОСФ"] if osf_row is not None else "—",
                "Тип матчинга": mtype,
                "Уверенность": conf,
                "Зона": risk["зона_риска"],
                "proba": round(float(risk["proba"]), 3),
            })
            if osf_row is None:
                continue
            osf_name = osf_row["ОСФ"]
            # если на одну ОСФ пришло несколько видов спорта — берём наихудшую зону
            prev = osf_to_risk.get(osf_name)
            z_new = ZONE_ORDER.get(risk["зона_риска"], 3)
            if prev is None or z_new < ZONE_ORDER.get(prev["зона_риска"], 3) or (
                z_new == ZONE_ORDER.get(prev["зона_риска"], 3)
                and float(risk["proba"]) > float(prev["proba"])
            ):
                osf_to_risk[osf_name] = {
                    "зона_риска": risk["зона_риска"],
                    "proba": float(risk["proba"]),
                    "причина": risk["причина"],
                    "вид_модель": risk["Вид спорта"],
                }

    results = []
    for _, osf in osf_df.iterrows():
        osf_name = osf["ОСФ"]
        rating   = int(osf["Баллы"])
        done_c   = osf["Выполненные критерии"]
        not_done_c = osf["Невыполненные критерии"]
        is_high_rat = rating >= RATING_THRESHOLD

        risk = osf_to_risk.get(osf_name)

        if risk is None:
            # Нет данных модели
            zone_code = ZONE_CODE_NODATA
            zone  = NO_DATA_ZONE
            proba = np.nan
            reason = "Вид спорта отсутствует в прогнозной модели"
            if is_high_rat:
                priority = 4
                quadrant = "Приоритет 4 (высокий рейтинг + нет систематичности)"
                rec = "Поддерживать текущий уровень; мониторинг; плановые проверки"
            else:
                priority = 3
                quadrant = "Приоритет 3 (низкий рейтинг + нет систематичности)"
                rec = "Продолжать профилактику; приоритизировать работу по невыполненным критериям рейтинга; превентивная работа с федерацией и регионом"
            miss_str = not_done_c if not_done_c else "нет"
            justification = (f"Данных модели нет; баллы рейтинга: {rating}/100"
                             + (f"; невыполненные критерии: {miss_str}" if not_done_c else "; критерии рейтинга выполнены полностью"))
        else:
            zone_raw = risk["зона_риска"]
            # Нормализуем зону к текстовому коду
            if zone_raw in ("🔴 КРАСНАЯ", "RED"):
                zone_code = ZONE_CODE_RED
                zone = "🔴 RED"
            elif zone_raw in ("🟠 ОРАНЖЕВАЯ", "ORANGE"):
                zone_code = ZONE_CODE_ORANGE
                zone = "🟠 ORANGE"
            elif zone_raw in ("🟢 ЗЕЛЁНАЯ", "GREEN"):
                zone_code = ZONE_CODE_GREEN
                zone = "🟢 GREEN"
            else:
                zone_code = ZONE_CODE_NODATA
                zone = "⚪ NO_DATA"
            proba  = risk["proba"]
            reason = risk["причина"]
            is_systematic = zone_code in (ZONE_CODE_RED, ZONE_CODE_ORANGE)

            if not is_high_rat and is_systematic:
                priority = 1
                quadrant = "Приоритет 1 (низкий рейтинг + систематичность)"
                rec = "Приоритизировать работу по невыполненным критериям рейтинга"
            elif is_high_rat and is_systematic:
                priority = 2
                quadrant = "Приоритет 2 (высокий рейтинг + систематичность)"
                rec = "Усилить антидопинговое образование; беседы с заинтересованными сторонами; рассмотреть внесоревновательное тестирование"
            elif not is_high_rat and not is_systematic:
                priority = 3
                quadrant = "Приоритет 3 (низкий рейтинг + нет систематичности)"
                rec = "Продолжать профилактику; приоритизировать работу по невыполненным критериям рейтинга; превентивная работа с федерацией и регионом"
            else:
                priority = 4
                quadrant = "Приоритет 4 (высокий рейтинг + нет систематичности)"
                rec = "Поддерживать текущий уровень; мониторинг; плановые проверки"

            proba_str = f"{proba:.3f}" if proba == proba else "нет данных"
            miss_str = not_done_c if not_done_c else "нет"
            if is_systematic:
                justification = (f"Рисковый из-за {reason} (зона {zone_code}, proba: {proba_str}); "
                                 + (f"нехватка баллов по критериям: {miss_str}" if not_done_c else "критерии рейтинга выполнены полностью"))
            else:
                justification = (f"Низкий риск по модели (proba: {proba_str}); баллы рейтинга: {rating}/100"
                                 + (f"; недостающие критерии: {miss_str}" if not_done_c else "; все критерии выполнены"))

        results.append({
            "Вид спорта (ОСФ)": osf_name,
            "Приоритет": priority,
            "Квадрант": quadrant,
            "Зона риска": zone,
            "_zone_code": zone_code,
            "Оценка риска (proba)": round(proba, 3) if proba == proba else None,
            "Баллы рейтинга РУСАДА": rating,
            "Место в рейтинге": osf["Место"] if pd.notna(osf["Место"]) else "",
            "Выполненные критерии РУСАДА": done_c,
            "Невыполненные критерии РУСАДА": not_done_c,
            "Рекомендация": rec,
            "Обоснование": justification,
        })

    df_out = pd.DataFrame(results)

    # Сортировка: приоритет → zone (RED=1, ORANGE=2, GREEN=3, NO_DATA=4) → proba убыв.
    zone_sort = {ZONE_CODE_RED: 1, ZONE_CODE_ORANGE: 2, ZONE_CODE_GREEN: 3, ZONE_CODE_NODATA: 4}
    df_out["_zone_sort"] = df_out["_zone_code"].map(zone_sort).fillna(4)
    df_out["_p"] = df_out["Оценка риска (proba)"].fillna(-1)
    df_out = df_out.sort_values(["Приоритет", "_zone_sort", "_p"],
                                ascending=[True, True, False])
    df_out = df_out.drop(columns=["_zone_sort", "_p"]).reset_index(drop=True)
    df_out.insert(0, "№", range(1, len(df_out) + 1))

    # ── САМОПРОВЕРКА (инварианты) ──────────────────────────────────────────
    n_rating = len(osf_df)
    n_matrix = len(df_out)
    print(f"\n{'='*60}")
    print("САМОПРОВЕРКА (инварианты)")
    print(f"{'='*60}")
    print(f"1. N_рейтинг={n_rating}, строк в матрице={n_matrix}: {'✓' if n_matrix == n_rating else '✗ ОШИБКА'}")
    prio_sum = df_out["Приоритет"].count()
    print(f"2. Сумма строк по приоритетам={prio_sum} (должно быть {n_matrix}): {'✓' if prio_sum == n_matrix else '✗ ОШИБКА'}")
    zone_dist = df_out["_zone_code"].value_counts().to_dict()
    print(f"3. Распределение по зонам: {zone_dist}")
    prio_dist = df_out["Приоритет"].value_counts().sort_index().to_dict()
    print(f"   Распределение по приоритетам: {prio_dist}")
    # Проверка 4: RED/ORANGE не в приоритетах 3-4
    bad4 = df_out[df_out["_zone_code"].isin([ZONE_CODE_RED, ZONE_CODE_ORANGE]) & df_out["Приоритет"].isin([3, 4])]
    print(f"4. RED/ORANGE с приоритетами 3-4: {len(bad4)} (должно быть 0): {'✓' if len(bad4)==0 else '✗ ОШИБКА'}")
    # Проверка 5: GREEN не в приоритетах 1-2
    bad5 = df_out[df_out["_zone_code"] == ZONE_CODE_GREEN & df_out["Приоритет"].isin([1, 2])]
    print(f"5. GREEN с приоритетами 1-2: {len(bad5)} (должно быть 0): {'✓' if len(bad5)==0 else '✗ ОШИБКА'}")
    print(f"{'='*60}\n")

    audit_df = pd.DataFrame(audit)
    return df_out, audit_df


# ══════════════════════════════════════════════════════════════════
# ШАГ 5. EXCEL С ФОРМАТИРОВАНИЕМ
# ══════════════════════════════════════════════════════════════════

def save_excel(df, audit_df, path):
    # Экспортируем без служебных колонок
    export_cols = [c for c in df.columns if not c.startswith("_")]
    df_export = df[export_cols].copy()

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df_export.to_excel(xl, index=False, sheet_name="Итог")
        if audit_df is not None and not audit_df.empty:
            audit_df.to_excel(xl, index=False, sheet_name="Аудит матчинга")

    wb = openpyxl.load_workbook(path)
    ws = wb["Итог"]

    # Цвета по ZONE_COLORS (текстовые коды)
    zone_fills = {
        ZONE_CODE_RED:    PatternFill("solid", fgColor="FECACA"),
        ZONE_CODE_ORANGE: PatternFill("solid", fgColor="FDE68A"),
        ZONE_CODE_GREEN:  PatternFill("solid", fgColor="BBF7D0"),
        ZONE_CODE_NODATA: PatternFill("solid", fgColor="E5E7EB"),
    }
    quad_fills = {
        "Приоритет 1 (низкий рейтинг + систематичность)": PatternFill("solid", fgColor="FECACA"),
        "Приоритет 2 (высокий рейтинг + систематичность)": PatternFill("solid", fgColor="FDE68A"),
        "Приоритет 3 (низкий рейтинг + нет систематичности)": PatternFill("solid", fgColor="BFDBFE"),
        "Приоритет 4 (высокий рейтинг + нет систематичности)": PatternFill("solid", fgColor="BBF7D0"),
    }

    header_fill = PatternFill("solid", fgColor="0F2D52")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = {1: 5, 2: 45, 3: 30, 4: 16, 5: 14, 6: 12, 7: 12,
                  8: 38, 9: 38, 10: 52, 11: 62}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    zone_col_idx = df_export.columns.get_loc("Зона риска") + 2 if "Зона риска" in df_export.columns else None
    quad_col_idx = df_export.columns.get_loc("Квадрант") + 2 if "Квадрант" in df_export.columns else None
    zone_code_col_idx = df.columns.get_loc("_zone_code") + 1 if "_zone_code" in df.columns else None

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        # Получаем zone_code из оригинального df
        zone_code_val = df["_zone_code"].iloc[row_idx - 1] if zone_code_col_idx and row_idx <= len(df) else ZONE_CODE_NODATA
        quad_val = row[quad_col_idx - 1].value if quad_col_idx else ""
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        if zone_col_idx:
            row[zone_col_idx - 1].fill = zone_fills.get(zone_code_val, PatternFill())
        if quad_col_idx:
            row[quad_col_idx - 1].fill = quad_fills.get(quad_val, PatternFill())

    ws.row_dimensions[1].height = 42
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 52
    ws.freeze_panes = "B2"

    wb.save(path)
    print(f"✅ Excel сохранён: {path}")


# ══════════════════════════════════════════════════════════════════
# ШАГ 6. PLOTLY ДАШБОРД
# ══════════════════════════════════════════════════════════════════

def _short_name(s):
    s = str(s)
    for prefix in ["Всероссийская федерация ", "Всероссийское спортивно-кинологическое объединение",
                   "Федерация ", "Российская Федерация ", "Российская федерация ",
                   "Российский ", "Союз ", "Ассоциация ", "Национальная федерация ",
                   "Национальный ", "Общероссийская спортивная федерация ",
                   "Объединенная федерация ", "Спортивная Федерация "]:
        if s.startswith(prefix):
            s = s[len(prefix):]
            break
    return s.replace(" России", "").replace("\n", " ").strip()


def build_dashboard(df, path):
    df_all = df.copy()
    df_all["_short"] = df_all["Вид спорта (ОСФ)"].apply(_short_name)

    df_risk = df_all[df_all["_zone_code"].isin([ZONE_CODE_RED, ZONE_CODE_ORANGE])].copy()

    # ── ГРАФИК 1: Матрица квадрантов (scatter) ──
    quad_order = list(QUADRANT_COLORS.keys())

    fig2 = go.Figure()
    rng = np.random.default_rng(42)
    for quad in quad_order:
        sub = df_all[df_all["Квадрант (риск × рейтинг)"] == quad].copy()
        if sub.empty:
            continue
        # для ОСФ без данных модели откладываем y около -0.05 (зона "нет данных")
        y_raw = sub["Оценка риска (proba)"].astype(float)
        y = y_raw.fillna(-0.05).values
        x = sub["Баллы рейтинга РУСАДА"].astype(float).values
        # лёгкий jitter, чтобы точки не сливались
        x = x + rng.uniform(-1.5, 1.5, size=len(x))
        y = y + rng.uniform(-0.012, 0.012, size=len(y))
        fig2.add_trace(go.Scatter(
            x=x, y=y, mode="markers", name=quad,
            marker=dict(color=QUADRANT_COLORS.get(quad, "#94A3B8"), size=12,
                        opacity=0.82, line=dict(width=1, color="white")),
            customdata=np.stack([
                sub["Вид спорта (ОСФ)"].values,
                sub["Зона риска (модель)"].values,
                sub["Обоснование"].values,
            ], axis=-1),
            hovertemplate=("<b>%{customdata[0]}</b><br>"
                           "Рейтинг: %{x:.0f} баллов<br>"
                           "Зона: %{customdata[1]}<br>"
                           "<i>%{customdata[2]}</i><extra></extra>"),
        ))

    fig2.add_vline(x=RATING_THRESHOLD, line_dash="dot", line_color="#94A3B8", line_width=1.5,
                   annotation_text=f"порог рейтинга = {RATING_THRESHOLD}",
                   annotation_position="top left",
                   annotation_font=dict(color="#94A3B8", size=10))
    fig2.add_hline(y=0.0, line_dash="dot", line_color="#CBD5E1", line_width=1)

    _theme(fig2, "Матрица Риск × Рейтинг ОСФ 2025",
           "ось X — баллы рейтинга РУСАДА · ось Y — оценка риска модели (⚪ нет данных ≈ ниже 0)")
    fig2.update_layout(
        height=700,
        xaxis=dict(range=[0, 110], title="Баллы рейтинга РУСАДА", showgrid=True, gridcolor=GRID),
        yaxis=dict(range=[-0.12, 1.05], title="Оценка риска (proba)", showgrid=True, gridcolor=GRID),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )

    # ── ГРАФИК 2: Топ рисковых видов спорта (hbar) ──
    if not df_risk.empty:
        d1 = df_risk.sort_values("Оценка риска (proba)").tail(20).copy()
        fig1 = px.bar(
            d1, x="Оценка риска (proba)", y="_short", orientation="h",
            color="_zone_code", color_discrete_map=ZONE_COLORS,
            text=d1["Оценка риска (proba)"].map(lambda v: f"{v:.2f}"),
            hover_name="Вид спорта (ОСФ)",
            hover_data={"Баллы рейтинга РУСАДА": True, "Квадрант (риск × рейтинг)": True,
                        "_short": False},
        )
        fig1.update_traces(marker_cornerradius=7, textposition="outside",
                           cliponaxis=False, textfont=dict(size=11, color=INK))
        _theme(fig1, "Рисковые виды спорта (ОСФ)",
               "цвет = зона: 🔴 свежая динамика · 🟠 история · топ-20 по оценке модели")
        fig1.update_layout(height=max(420, 32 * len(d1)), yaxis_title="",
                           xaxis_title="оценка риска (0–1)", legend_title_text="зона")
        fig1.update_xaxes(showgrid=True, gridcolor=GRID, range=[0, 1.15])
    else:
        fig1 = go.Figure()
        _theme(fig1, "Рисковые виды спорта (ОСФ)", "нет рисковых зон в данных модели")
        fig1.update_layout(height=300)

    # ── ГРАФИК 3: Распределение по квадрантам (pie) ──
    quad_counts = df_all["Квадрант"].value_counts().reset_index()
    quad_counts.columns = ["Квадрант", "Кол-во"]
    fig3 = go.Figure(go.Pie(
        labels=quad_counts["Квадрант"], values=quad_counts["Кол-во"],
        marker_colors=[QUADRANT_COLORS.get(q, "#94A3B8") for q in quad_counts["Квадрант"]],
        textinfo="label+percent+value",
        hovertemplate="<b>%{label}</b><br>Федераций: %{value}<br>%{percent}<extra></extra>",
        hole=0.4, textfont=dict(size=11),
    ))
    _theme(fig3, "Распределение ОСФ по квадрантам", f"всего федераций: {len(df_all)}")
    fig3.update_layout(height=480, showlegend=False)

    # ── СБОРКА HTML ──
    parts = ["""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Антидопинговый дашборд: Риск × Рейтинг ОСФ 2025</title>
<style>
  body { font-family: Inter, Segoe UI, Roboto, Arial, sans-serif; background:#F8FAFC; color:#0F2D52; margin:0; }
  .header { background:#0F2D52; color:white; padding:24px 32px; }
  .header h1 { margin:0; font-size:22px; }
  .header p { margin:6px 0 0; font-size:13px; color:#7DD3FC; }
  .section { padding:24px 32px; }
  .section h2 { font-size:16px; border-left:4px solid #2563EB; padding-left:10px; margin-bottom:4px; }
  .section p { font-size:12px; color:#7C8DA6; margin:0 0 12px; }
  .divider { border:none; border-top:1px solid #EEF2F7; margin:0 32px; }
</style></head><body>
<div class="header">
  <h1>Антидопинговый дашборд: Матрица Риск × Рейтинг ОСФ 2025</h1>
  <p>Для работников антидопинговой сферы · Источники: Рейтинг ОСФ РУСАДА 2025 + прогнозная модель нарушений</p>
</div>"""]

    parts.append('<div class="section"><h2>Матрица Риск × Рейтинг</h2>'
                 '<p>Каждая точка — ОСФ. ⚪ серые точки внизу — федерации без данных модели. Наведите для деталей.</p>')
    parts.append(fig2.to_html(full_html=False, include_plotlyjs="cdn"))
    parts.append("</div><hr class='divider'>")

    parts.append('<div class="section"><h2>Топ рисковых видов спорта</h2>'
                 '<p>Только зоны 🔴 и 🟠 · сортировка по оценке модели</p>')
    parts.append(fig1.to_html(full_html=False, include_plotlyjs=False))
    parts.append("</div><hr class='divider'>")

    parts.append('<div class="section"><h2>Распределение по квадрантам</h2>'
                 '<p>Доля ОСФ в каждом квадранте матрицы</p>')
    parts.append(fig3.to_html(full_html=False, include_plotlyjs=False))
    parts.append("</div></body></html>")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))
    print(f"✅ HTML-дашборд сохранён: {path}")


# ══════════════════════════════════════════════════════════════════
# ЗАПУСК
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("📂 Загружаем рейтинг ОСФ из PDF...")
    osf_df = load_osf_rating(PDF_PATH)
    print(f"   ОСФ в рейтинге: {len(osf_df)}")

    if os.path.exists(IPYNB_PATH):
        risks_df = load_risks_from_ipynb(IPYNB_PATH)
    else:
        print(f"   ⚠ ipynb не найден ({IPYNB_PATH}) — риски будут помечены как «нет данных»")
        risks_df = pd.DataFrame()
    print(f"   Видов спорта в модели: {len(risks_df)}")

    print("🔗 Матчинг и сборка таблицы по всем ОСФ...")
    df_final, audit_df = build_final_table(osf_df, risks_df)
    print(f"   Итоговых строк (= всех ОСФ): {len(df_final)}")
    for z in [ZONE_CODE_RED, ZONE_CODE_ORANGE, ZONE_CODE_GREEN, ZONE_CODE_NODATA]:
        print(f"   {ZONE_EMOJI.get(z,'')} {z}: {(df_final['_zone_code'] == z).sum()}")

    excel_path = os.path.join(OUT_DIR, "osf_risk_final.xlsx")
    html_path  = os.path.join(OUT_DIR, "osf_risk_dashboard_final.html")

    print("💾 Сохраняем Excel...")
    save_excel(df_final, audit_df, excel_path)

    print("📊 Строим дашборд...")
    build_dashboard(df_final, html_path)

    print(f"\n✅ Готово! Файлы в: {OUT_DIR}")
